import os   # To read the files
import openpyxl     # To read/write to excel sheet

wb = openpyxl.load_workbook(filename = 'Pokemon Info.xlsx')
sheet = wb['File Checker']

def find_col_num(s):
    # Because row/column numbers in openpyxl start at 1
    for i in range(1, sheet.max_column):
        if sheet.cell(1, i).value == s:
            return (i)

files = os.listdir("C:\\Users\\ejone\\OneDrive\\Desktop\\Code\\Javascript\\p5\\projects\\Pokeball Pokemon Comparison\\Images\\Pokemon\\Game Sprites")
print(files)

# TODO: Going to need to do for regular files (ending in animated, static, or shiny)
    # And forms
        # Will have to add form rows below pokemon to excel sheet