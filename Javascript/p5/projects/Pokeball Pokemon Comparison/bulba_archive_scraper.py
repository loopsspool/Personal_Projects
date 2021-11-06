import requests # To retrieve webpages
from bs4 import BeautifulSoup   # To parse webpages
import re   # For filtering what images to download
import urllib   # For downloading those images to my computer
import os   # For downloading those images to my computer
from PIL import Image   # For converting URL image data to PIL Image object 
import openpyxl     # For reading excel workbook
# Must explicitly state this...
from openpyxl import load_workbook
import string # To access letters easily without having to type them myself in an array

# TODO: Tidy up this damn file... smh

# NOTE: ALL DOWNLOADS MUST BE DONE IN THE FASHION BELOW
    # Otherwise bulba has a check on if the site is being web scraped and it will block the download
# This is to mask the fact I'm webscraping
    # To use, call
        # filename, headers = opener.retrieve(get_largest_png(img), gen8_menu_sprite_save_path + save_name)
opener = urllib.request.URLopener()
opener.addheader('User-Agent', 'Mozilla/5.0')

def search_for_drawn_forms(pokemon):
    # Custom type forms
    # Pikachu Cosplay & Caps
    get_img_from_string(img, "^\d\d\dPikachu-Alola.png", drawn_save_path + save_name + "-Cap-Alola")
    get_img_from_string(img, "^\d\d\dPikachu-Hoenn.png", drawn_save_path + save_name + "-Cap-Hoenn")
    get_img_from_string(img, "^\d\d\dPikachu-Kalos.png", drawn_save_path + save_name + "-Cap-Kalos")
    get_img_from_string(img, "^\d\d\dPikachu-Original.png", drawn_save_path + save_name + "-Cap-Original")
    get_img_from_string(img, "^\d\d\dPikachu-Partner.png", drawn_save_path + save_name + "-Cap-Partner")
    get_img_from_string(img, "^\d\d\dPikachu-Sinnoh.png", drawn_save_path + save_name + "-Cap-Sinnoh")
    get_img_from_string(img, "^\d\d\dPikachu-Unova.png", drawn_save_path + save_name + "-Cap-Unova")
    get_img_from_string(img, "^\d\d\dPikachu-World.png", drawn_save_path + save_name + "-Cap-World")
    get_img_from_string(img, "^\d\d\dPikachu-Belle.png", drawn_save_path + save_name + "-Cosplay-Belle")
    get_img_from_string(img, "^\d\d\dPikachu-Libre.png", drawn_save_path + save_name + "-Cosplay-Libre")
    get_img_from_string(img, "^\d\d\dPikachu-PhD.png", drawn_save_path + save_name + "-Cosplay-PhD")
    get_img_from_string(img, "^\d\d\dPikachu-Pop Star.png", drawn_save_path + save_name + "-Cosplay-Pop_Star")
    get_img_from_string(img, "^\d\d\dPikachu-Rock Star.png", drawn_save_path + save_name + "-Cosplay-Rock_Star")

    # Spiky-eared Pichu
    get_img_from_string(img, "Spiky-eared Pichu DP 1", drawn_save_path + save_name + "-Spiky_Eared")

    # Unown Characters
    if pokemon.name == "Unown":
        # Only drawn forms are dream versions
        if img["alt"].endswith("Dream.png"):
            # Get form
            form = img["alt"].split(" ")[1]
            if form == "Exclamation":
                form = "!"    
            if form == "Question":
                form = "Qmark"
            form = "-" + form
            get_img_from_string(img, "^\d\d\dUnown [a-zA-z]+ Dream.png", drawn_save_path + save_name + form)

    # Castform Weathers
    get_img_from_string(img, "^\d\d\dCastform-Rainy.png", drawn_save_path + save_name + "-Rainy")
    get_img_from_string(img, "^\d\d\dCastform-Snowy.png", drawn_save_path + save_name + "-Snowy")
    get_img_from_string(img, "^\d\d\dCastform-Sunny.png", drawn_save_path + save_name + "-Sunny")

    # Primal Kyogre & Groudon
    get_img_from_string(img, "^\d\d\dKyogre-Primal 2.png", drawn_save_path + save_name + "-Primal")
    get_img_from_string(img, "^\d\d\dGroudon-Primal.png", drawn_save_path + save_name + "-Primal")

    # Deoxys
    get_img_from_string(img, "^\d\d\dDeoxys-Attack.png", drawn_save_path + save_name + "-Attack")
    get_img_from_string(img, "^\d\d\dDeoxys-Defense.png", drawn_save_path + save_name + "-Defense")
    get_img_from_string(img, "^\d\d\dDeoxys-Speed.png", drawn_save_path + save_name + "-Speed")

    # Burmy & Wormadam Cloaks
    get_img_from_string(img, "^\d\d\dBurmy-Plant.png", drawn_save_path + save_name + "-Plant")
    get_img_from_string(img, "^\d\d\dBurmy-Sandy.png", drawn_save_path + save_name + "-Sandy")
    get_img_from_string(img, "^\d\d\dBurmy-Trash.png", drawn_save_path + save_name + "-Trash")
    get_img_from_string(img, "^\d\d\dWormadam-Plant.png", drawn_save_path + save_name + "-Plant")
    get_img_from_string(img, "^\d\d\dWormadam-Sandy.png", drawn_save_path + save_name + "-Sandy")
    get_img_from_string(img, "^\d\d\dWormadam-Trash.png", drawn_save_path + save_name + "-Trash")

    # Cherrim
    # NOTE: No default image, only overcast and sunny
    get_img_from_string(img, "^\d\d\dCherrim-Overcast.png", drawn_save_path + save_name + "-Overcast")
    get_img_from_string(img, "^\d\d\dCherrim-Sunny.png", drawn_save_path + save_name + "-Sunshine")

    # Shellos & Gastrodon East/West
    # NOTE: No default image
    get_img_from_string(img, "^\d\d\dShellos-East.png", drawn_save_path + save_name + "-East")
    get_img_from_string(img, "^\d\d\dShellos-West.png", drawn_save_path + save_name + "-West")
    get_img_from_string(img, "^\d\d\dGastrodon-East.png", drawn_save_path + save_name + "-East")
    get_img_from_string(img, "^\d\d\dGastrodon-West.png", drawn_save_path + save_name + "-West")

    # Rotom Appliances
    get_img_from_string(img, "^\d\d\dRotom-Fan.png", drawn_save_path + save_name + "-Fan")
    get_img_from_string(img, "^\d\d\dRotom-Frost.png", drawn_save_path + save_name + "-Frost")
    get_img_from_string(img, "^\d\d\dRotom-Heat.png", drawn_save_path + save_name + "-Heat")
    get_img_from_string(img, "^\d\d\dRotom-Mow.png", drawn_save_path + save_name + "-Mow")
    get_img_from_string(img, "^\d\d\dRotom-Wash.png", drawn_save_path + save_name + "-Wash")

    # Giratina
    # NOTE: No default image
    get_img_from_string(img, "^\d\d\dGiratina-Altered.png", drawn_save_path + save_name + "-Altered")
    get_img_from_string(img, "^\d\d\dGiratina-Origin.png", drawn_save_path + save_name + "-Origin")

    # Shaymin
    # NOTE: No default image
    get_img_from_string(img, "^\d\d\dShaymin-Land.png", drawn_save_path + save_name + "-Land")
    get_img_from_string(img, "^\d\d\dShaymin-Sky.png", drawn_save_path + save_name + "-Sky")

    # Arceus Types
    # Only drawn forms are dream versions
    if pokemon.name == "Arceus":
        if img["alt"].endswith("Dream.png"):
            # Get form
            form = img["alt"].split(" ")[1]
            form = "-" + form
            get_img_from_string(img, "^\d\d\dArceus [a-zA-z]+ Dream.png", drawn_save_path + save_name + form)

    # Basculin Stripes
    get_img_from_string(img, "^\d\d\dBasculin-Red-Striped_XY_Anime.png", drawn_save_path + save_name + "-Red_Striped")
    get_img_from_string(img, "^\d\d\dBasculin-Blue-Striped_BW_Anime.png", drawn_save_path + save_name + "-Blue_Striped")

    # Darmanitan Modes
    get_img_from_string(img, "^\d\d\dDarmanitan.png", drawn_save_path + save_name + "-Standard")
    get_img_from_string(img, "^\d\d\dDarmanitan-Galar.png", drawn_save_path + save_name + "-Region-Galar-Standard")
    get_img_from_string(img, "^\d\d\dDarmanitan-Zen.png", drawn_save_path + save_name + "-Zen")
    get_img_from_string(img, "^\d\d\dDarmanitan-Galar-Zen.png", drawn_save_path + save_name + "-Region-Galar-Zen")

    # Deerling & Sawsbuck Seasons
    # NOTE: No default image
    get_img_from_string(img, "^\d\d\dDeerling-Autumn.png", drawn_save_path + save_name + "-Autumn")
    get_img_from_string(img, "^\d\d\dDeerling-Spring.png", drawn_save_path + save_name + "-Spring")
    get_img_from_string(img, "^\d\d\dDeerling-Summer.png", drawn_save_path + save_name + "-Summer")
    get_img_from_string(img, "^\d\d\dDeerling-Winter.png", drawn_save_path + save_name + "-Winter")
    get_img_from_string(img, "^\d\d\dSawsbuck-Autumn.png", drawn_save_path + save_name + "-Autumn")
    get_img_from_string(img, "^\d\d\dSawsbuck-Spring.png", drawn_save_path + save_name + "-Spring")
    get_img_from_string(img, "^\d\d\dSawsbuck-Summer.png", drawn_save_path + save_name + "-Summer")
    get_img_from_string(img, "^\d\d\dSawsbuck-Winter.png", drawn_save_path + save_name + "-Winter")

    # Forces of nature forms
    get_img_from_string(img, "^\d\d\dTornadus.png", drawn_save_path + save_name + "-Incarnate")
    get_img_from_string(img, "^\d\d\dTornadus-Therian.png", drawn_save_path + save_name + "-Therian")
    get_img_from_string(img, "^\d\d\dThundurus.png", drawn_save_path + save_name + "-Incarnate")
    get_img_from_string(img, "^\d\d\dThundurus-Therian.png", drawn_save_path + save_name + "-Therian")
    get_img_from_string(img, "^\d\d\dLandorus.png", drawn_save_path + save_name + "-Incarnate")
    get_img_from_string(img, "^\d\d\dLandorus-Therian.png", drawn_save_path + save_name + "-Therian")

    # Kyurem Fusions
    get_img_from_string(img, "^\d\d\dKyurem-Black.png", drawn_save_path + save_name + "-Black")
    get_img_from_string(img, "^\d\d\dKyurem-Black2.png", drawn_save_path + save_name + "-Black_Overdrive")
    get_img_from_string(img, "^\d\d\dKyurem-White.png", drawn_save_path + save_name + "-White")
    get_img_from_string(img, "^\d\d\dKyurem-White2.png", drawn_save_path + save_name + "-White_Overdrive")
    
    # Keldeo
    get_img_from_string(img, "^\d\d\dKeldeo.png", drawn_save_path + save_name + "-Ordinary")
    get_img_from_string(img, "^\d\d\dKeldeo-Resolute.png", drawn_save_path + save_name + "-Resolute")

    # Meloetta
    get_img_from_string(img, "^\d\d\dMeloetta.png", drawn_save_path + save_name + "-Aria")
    get_img_from_string(img, "^\d\d\dMeloetta-Pirouette.png", drawn_save_path + save_name + "-Pirouette")

    # Genesect
    # Only drawn forms are dream versions
    if pokemon.name == "Genesect":
        if img["alt"].endswith("Dream.png"):
            # Get form
            form = img["alt"].split(" ")[1]
            if form == "B":
                form = "Burn_Drive"
            if form == "C":
                form = "Chill_Drive"
            if form == "D":
                form = "Douse_Drive"
            if form == "S":
                form = "Shock_Drive"
            form = "-" + form
            get_img_from_string(img, "^\d\d\dGenesect [a-zA-z] Dream.png", drawn_save_path + save_name + form)

    # Ash Greninja
    get_img_from_string(img, "^\d\d\dGreninja-Ash.png", drawn_save_path + save_name + "-Ash")

    # Vivillon Patterns
    get_img_from_string(img, "^\d\d\dVivillon-Archipelago.png", drawn_save_path + save_name + "-Archipelago")
    get_img_from_string(img, "^\d\d\dVivillon-Continental.png", drawn_save_path + save_name + "-Continental")
    get_img_from_string(img, "^\d\d\dVivillon-Elegant.png", drawn_save_path + save_name + "-Elegant")
    get_img_from_string(img, "^\d\d\dVivillon-Fancy.png", drawn_save_path + save_name + "-Fancy")
    get_img_from_string(img, "^\d\d\dVivillon-Garden.png", drawn_save_path + save_name + "-Garden")
    get_img_from_string(img, "^\d\d\dVivillon-High Plains.png", drawn_save_path + save_name + "-High_Plains")
    get_img_from_string(img, "^\d\d\dVivillon-Icy Snow.png", drawn_save_path + save_name + "-Icy_Snow")
    get_img_from_string(img, "^\d\d\dVivillon-Jungle.png", drawn_save_path + save_name + "-Jungle")
    get_img_from_string(img, "^\d\d\dVivillon-Marine.png", drawn_save_path + save_name + "-Marine")
    get_img_from_string(img, "^\d\d\dVivillon-Meadow.png", drawn_save_path + save_name + "-Meadow")
    get_img_from_string(img, "^\d\d\dVivillon-Modern.png", drawn_save_path + save_name + "-Modern")
    get_img_from_string(img, "^\d\d\dVivillon-Monsoon.png", drawn_save_path + save_name + "-Monsoon")
    get_img_from_string(img, "^\d\d\dVivillon-Ocean.png", drawn_save_path + save_name + "-Ocean")
    get_img_from_string(img, "^\d\d\dVivillon-Poké Ball.png", drawn_save_path + save_name + "-Poke_Ball")
    get_img_from_string(img, "^\d\d\dVivillon-Polar.png", drawn_save_path + save_name + "-Polar")
    get_img_from_string(img, "^\d\d\dVivillon-River.png", drawn_save_path + save_name + "-River")
    get_img_from_string(img, "^\d\d\dVivillon-Sandstorm.png", drawn_save_path + save_name + "-Sandstorm")
    get_img_from_string(img, "^\d\d\dVivillon-Savanna.png", drawn_save_path + save_name + "-Savanna")
    get_img_from_string(img, "^\d\d\dVivillon-Sun.png", drawn_save_path + save_name + "-Sun")
    get_img_from_string(img, "^\d\d\dVivillon-Tundra.png", drawn_save_path + save_name + "-Tundra")


    # Flabebe, Floette, and Florges colors
    get_img_from_string(img, "^\d\d\dFlabébé Blue Flower XY anime.png", drawn_save_path + save_name + "-Blue")
    get_img_from_string(img, "^\d\d\dFlabébé Orange Flower XY anime.png", drawn_save_path + save_name + "-Orange")
    get_img_from_string(img, "^\d\d\dFlabébé Red Flower XY anime.png", drawn_save_path + save_name + "-Red")
    get_img_from_string(img, "^\d\d\dFlabébé White Flower XY anime.png", drawn_save_path + save_name + "-White")
    get_img_from_string(img, "^\d\d\dFlabébé Yellow Flower XY anime.png", drawn_save_path + save_name + "-Yellow")
    get_img_from_string(img, "^\d\d\dFloette-Blue XY anime.png", drawn_save_path + save_name + "-Blue")
    get_img_from_string(img, "^\d\d\dFloette-Orange XY anime.png", drawn_save_path + save_name + "-Orange")
    get_img_from_string(img, "^\d\d\dFloette-Red XY anime.png", drawn_save_path + save_name + "-Red")
    get_img_from_string(img, "^\d\d\dFloette-Yellow XY anime.png", drawn_save_path + save_name + "-Yellow")
    #get_img_from_string(img, "^\d\d\d.png", drawn_save_path + save_name + )
    get_img_from_string(img, "^\d\d\dFlorges Blue Flower XY anime.png", drawn_save_path + save_name + "-Blue")
    get_img_from_string(img, "^\d\d\dFlorges Orange Flower XY anime.png", drawn_save_path + save_name + "-Orange")
    get_img_from_string(img, "^\d\d\dFlorges Red Flower XY anime.png", drawn_save_path + save_name + "-Red")
    get_img_from_string(img, "^\d\d\dFlorges White Flower XY anime.png", drawn_save_path + save_name + "-White")
    get_img_from_string(img, "^\d\d\dFlorges Yellow Flower XY anime.png", drawn_save_path + save_name + "-Yellow")

    # Furfrou Trims
    get_img_from_string(img, "^\d\d\dFurfrou-Diamond.png", drawn_save_path + save_name + "-Diamond_Trim")
    get_img_from_string(img, "^\d\d\dFurfrou-Heart.png", drawn_save_path + save_name + "-Heart_Trim")
    get_img_from_string(img, "^\d\d\dFurfrou-Star.png", drawn_save_path + save_name + "-Star_Trim")
    # get_img_from_string(img, "^\d\d\d.png", drawn_save_path + save_name + )
    # get_img_from_string(img, "^\d\d\d.png", drawn_save_path + save_name + )
    # get_img_from_string(img, "^\d\d\d.png", drawn_save_path + save_name + )
    # get_img_from_string(img, "^\d\d\d.png", drawn_save_path + save_name + )
    # get_img_from_string(img, "^\d\d\d.png", drawn_save_path + save_name + )
    # get_img_from_string(img, "^\d\d\d.png", drawn_save_path + save_name + )

    # Aegislash
    get_img_from_string(img, "^\d\d\dAegislash-Blade.png", drawn_save_path + save_name + "-Blade")
    get_img_from_string(img, "^\d\d\dAegislash-Shield.png", drawn_save_path + save_name + "-Shield")

    # Pumpkaboo and Gourgeist Sizes
    # if "Pumpkaboo" == split_name or "Gourgeist" == split_name:
    #     # Average sizes have no indication in filename on this website
    #     form = " 1Average Size"
    # else:
    #     get_img_from_string(img, "^\d\d\d.png", drawn_save_path + save_name + )
    #     get_img_from_string(img, "^\d\d\d.png", drawn_save_path + save_name + )
    #     get_img_from_string(img, "^\d\d\d.png", drawn_save_path + save_name + )
    #     get_img_from_string(img, "^\d\d\d.png", drawn_save_path + save_name + )
    #     get_img_from_string(img, "^\d\d\d.png", drawn_save_path + save_name + )
    #     get_img_from_string(img, "^\d\d\d.png", drawn_save_path + save_name + )

    # Xerneas
    # if "Xerneas" == split_name:
    #     form = " Active"
    # else:
    #     get_img_from_string(img, "^\d\d\d.png", drawn_save_path + save_name + )

    # Zygarde
    get_img_from_string(img, "^\d\d\dZygarde.png", drawn_save_path + save_name + "-50%")
    get_img_from_string(img, "^\d\d\dZygarde-10Percent.png", drawn_save_path + save_name + "-10%")
    get_img_from_string(img, "^\d\d\dZygarde-Complete.png", drawn_save_path + save_name + "-Complete")


    # Hoopa
    get_img_from_string(img, "^\d\d\dHoopa.png", drawn_save_path + save_name + "-Confined")
    get_img_from_string(img, "^\d\d\dHoopa-Unbound.png", drawn_save_path + save_name + "-Unbound")


    # Oricorio
    # NOTE: No default
    get_img_from_string(img, "^\d\d\dOricorio-Baile.png", drawn_save_path + save_name + "-Baile")
    get_img_from_string(img, "^\d\d\dOricorio-Pa'u.png", drawn_save_path + save_name + "-Pa'u")
    get_img_from_string(img, "^\d\d\dOricorio-Pom-Pom.png", drawn_save_path + save_name + "-Pom_Pom")
    get_img_from_string(img, "^\d\d\dOricorio-Sensu.png", drawn_save_path + save_name + "-Sensu")

    # Lycanroc
    get_img_from_string(img, "^\d\d\dLycanroc.png", drawn_save_path + save_name + "-Midday")
    get_img_from_string(img, "^\d\d\dLycanroc-Dusk.png", drawn_save_path + save_name + "-Dusk")
    get_img_from_string(img, "^\d\d\dLycanroc-Midnight.png", drawn_save_path + save_name + "-Midnight")

    # Wishiwashi
    # NOTE: No default
    get_img_from_string(img, "^\d\d\dWishiwashi-Solo.png", drawn_save_path + save_name + "-Solo")
    get_img_from_string(img, "^\d\d\dWishiwashi-School.png", drawn_save_path + save_name + "-School")

    # Silvally Types
    # Only drawn forms are dream versions
    if pokemon.name == "Silvally":
        if img["alt"].endswith("Dream.png"):
            # Get form
            form = img["alt"].split(" ")[1]
            form = "-" + form
            get_img_from_string(img, "^\d\d\dSilvally [a-zA-z]+ Dream.png", drawn_save_path + save_name + form)

    # Minior
    get_img_from_string(img, "^\d\d\dMinior.png", drawn_save_path + save_name + "-Meteor")
    get_img_from_string(img, "^\d\d\dMinior-Core.png", drawn_save_path + save_name + "-Red_Core")
    # get_img_from_string(img, "^\d\d\d.png", drawn_save_path + save_name + )
    # get_img_from_string(img, "^\d\d\d.png", drawn_save_path + save_name + )
    # get_img_from_string(img, "^\d\d\d.png", drawn_save_path + save_name + )
    # get_img_from_string(img, "^\d\d\d.png", drawn_save_path + save_name + )
    # get_img_from_string(img, "^\d\d\d.png", drawn_save_path + save_name + )
    # get_img_from_string(img, "^\d\d\d.png", drawn_save_path + save_name + )
    # get_img_from_string(img, "^\d\d\d.png", drawn_save_path + save_name + )
    # get_img_from_string(img, "^\d\d\d.png", drawn_save_path + save_name + )
    # # Shiny cores all the same color?
    # get_img_from_string(img, "^\d\d\d.png", drawn_save_path + save_name + )

    # Mimikyu
    get_img_from_string(img, "^\d\d\dMimikyu.png", drawn_save_path + save_name + "-Disguised")
    get_img_from_string(img, "^\d\d\dMimikyu Busted Dream.png", drawn_save_path + save_name + "-Busted")

    # Solgaleo
    get_img_from_string(img, "^\d\d\dSolgaleo-RadiantSunPhase.png", drawn_save_path + save_name + "-Radiant_Sun")

    # Lunala
    get_img_from_string(img, "^\d\d\dLunala-FullMoonPhase.png", drawn_save_path + save_name + "-Full_Moon")

    # Necrozma
    get_img_from_string(img, "^\d\d\dNecrozma-Dawn Wings.png", drawn_save_path + save_name + "-Dawn_Wings")
    get_img_from_string(img, "^\d\d\dNecrozma-Dusk Mane.png", drawn_save_path + save_name + "-Dusk_Mane")
    get_img_from_string(img, "^\d\d\dNecrozma-Ultra.png", drawn_save_path + save_name + "-Ultra")

    # Magearna
    # get_img_from_string(img, "^\d\d\d.png", drawn_save_path + save_name + )

    # Marshadow
    get_img_from_string(img, "^\d\d\dMarshadow-Alt.png", drawn_save_path + save_name + "-Zenith")

    # Cramorant
    get_img_from_string(img, "^\d\d\dCramorant-Gorging.png", drawn_save_path + save_name + "-Gorging")
    get_img_from_string(img, "^\d\d\dCramorant-Gulping.png", drawn_save_path + save_name + "-Gulping")

    # Toxtricity
    get_img_from_string(img, "^\d\d\dToxtricity-Amped.png", drawn_save_path + save_name + "-Amped")
    get_img_from_string(img, "^\d\d\dToxtricity-Low Key.png", drawn_save_path + save_name + "-Low_Key")

    # Alcremie Creams & Sweets
    # Default Alcremie is Vanilla Cream-Strawberry Sweet
    if pokemon.name == "Alcremie":
        # Space after excludes gigantamax img
        if re.search("^869Alcremie-[a-zA-Z]+ ", img.attrs["alt"]):
            # Getting largest image for Alcremie
            img_url = get_largest_png(img)
            # Splits by directory
            img_url = img_url.split("/")
            # Gets last string in sequence (the filename)
            img_url = img_url[len(img_url) - 1]
            # Splits by hyphen to get cream and sweet
            img_url = img_url.split("-")
            cream = "-" + img_url[2]
            sweet = "-" + img_url[3].replace(".png 2x", "_Sweet")
            form = cream + sweet
            get_img_from_string(img, "^869Alcremie-[a-zA-Z]+ ", drawn_save_path + save_name + form)

    # Eiscue
    # NOTE: No default
    get_img_from_string(img, "^\d\d\dEiscue-Ice.png", drawn_save_path + save_name + "-Ice_Face")
    get_img_from_string(img, "^\d\d\dEiscue-Noice.png", drawn_save_path + save_name + "-Noice_Face")

    # Morpeko
    get_img_from_string(img, "^\d\d\dMorpeko-Full.png", drawn_save_path + save_name + "-Full")
    get_img_from_string(img, "^\d\d\dMorpeko-Hangry.png", drawn_save_path + save_name + "-Hangry")


    # Zacian and Zamazenta
    get_img_from_string(img, "^\d\d\dZacian.png", drawn_save_path + save_name + "-Crowned_Sword")
    get_img_from_string(img, "^\d\d\dZacian-Hero.png", drawn_save_path + save_name + "-Hero_of_Many_Battles")
    get_img_from_string(img, "^\d\d\dZamazenta.png", drawn_save_path + save_name + "-Crowned_Shield")
    get_img_from_string(img, "^\d\d\dZamazenta-Hero.png", drawn_save_path + save_name + "-Hero_of_Many_Battles")

    # Eternatus Eternamax
    # get_img_from_string(img, "^\d\d\d.png", drawn_save_path + save_name + )

    # Urshifu
    get_img_from_string(img, "^\d\d\dUrshifu-Gigantamax Rapid Strike.png", drawn_save_path + save_name + "Gigantamax-Rapid_Strike")
    get_img_from_string(img, "^\d\d\dUrshifu-Gigantamax Single Strike.png", drawn_save_path + save_name + "Gigantamax-Single_Strike")
    get_img_from_string(img, "^\d\d\dUrshifu-Rapid Strike.png", drawn_save_path + save_name + "-Rapid_Strike")
    get_img_from_string(img, "^\d\d\dUrshifu-Single Strike.png", drawn_save_path + save_name + "-Single_Strike")


    # Zarude
    get_img_from_string(img, "^\d\d\dZarude-Dada JN anime.png", drawn_save_path + save_name + "-Dada")

    # Calyrex Ridings
    get_img_from_string(img, "^\d\d\dCalyrex-Ice Rider.png", drawn_save_path + save_name + "-Ice_Rider")
    get_img_from_string(img, "^\d\d\dCalyrex-Shadow Rider.png", drawn_save_path + save_name + "-Shadow_Rider")

    # Zekrom Overdrive
    get_img_from_string(img, "^\d\d\dZekrom-Activated.png", drawn_save_path + save_name + "-Overdrive")
    # Reshiram Overdrive
    get_img_from_string(img, "^\d\d\dReshiram-Activated.png", drawn_save_path + save_name + "-Overdrive")


def get_drawn_images(pokemon, img):
    # DRAWN IMAGES
    # Drawn standard

    save_name = pokemon.number + " " + pokemon.name
    if pokemon.name == "Type: Null":
        save_name = pokemon.number + " Type Null"
    # Done this way so certain images that just have characters after the pokemon number don't match
        # Don't have to do this with the others because the hyphen denoters prevent the possibility
    pokemon_name_len = len(pokemon.name)
    get_img_from_string(img, "^\d\d\d[a-zA-Z]{" + str(pokemon_name_len) + "}.png", drawn_save_path + save_name)
    # Drawn Mega
    if pokemon.has_mega:
        if pokemon.name == "Charizard" or pokemon.name == "Mewtwo":
            get_img_from_string(img, "^\d\d\d[a-zA-Z]-Mega X.png", drawn_save_path + save_name + "-Mega_X")
            get_img_from_string(img, "^\d\d\d[a-zA-Z]-Mega Y.png", drawn_save_path + save_name + "-Mega_Y")
        else:
            get_img_from_string(img, "^\d\d\d[a-zA-Z]-Mega.png", drawn_save_path + save_name + "-Mega")
    # Gigantamax
    if pokemon.has_giganta:
        get_img_from_string(img, "^\d\d\d[a-zA-Z]-Gigantamax.png", drawn_save_path + save_name + "-Gigantamax")
    # Regional forms
    if pokemon.reg_forms != "":
        if "," in pokemon.reg_forms:
            get_img_from_string(img, "^\d\d\d[a-zA-Z]-Alola.png", drawn_save_path + save_name + "-Region-Alola")
            get_img_from_string(img, "^\d\d\d[a-zA-Z]-Galar.png", drawn_save_path + save_name + "-Region-Galar")
        else:
            get_img_from_string(img, "^\d\d\d[a-zA-Z]-" + pokemon.reg_forms + ".png", drawn_save_path + save_name + "-Region-" + pokemon.reg_forms)
    # Other forms
    if pokemon.has_misc_forms or pokemon.has_type_forms:
        search_for_drawn_forms(pokemon)

def get_menu_sprites():
    print("Getting Menu Sprites...")
    # session = requests.Session()
    # response = session.get(url, headers={'user-agent': 'Mozilla/5.0'})

    ms_end_urls = ["Generation_VI_menu_sprites", "Generation_VIII_menu_sprites"]
    for end_url in ms_end_urls:
        ms_page = requests.get("https://archives.bulbagarden.net/wiki/Category:" + end_url, headers={'User-Agent': 'Mozilla/5.0'})
        curr_ms_page_soup = BeautifulSoup(ms_page.content, 'html.parser')
        theres_a_next_page = True
        while (theres_a_next_page):
            pokemon_imgs = curr_ms_page_soup.find_all('img')
            # Downloading certain images
            for img in pokemon_imgs:
                img_text = img.attrs['alt']
                # If image doesn't match pokemon number + menu sprite + gen formula, don't bother
                if not re.search("\d\d\dMS\d", img_text):
                    continue
                poke_num = img_text[:3]
                poke_name = pokedex[int(poke_num) - 1].name
                file_ext = img_text[len(img_text) - 4:]
                # The zfill adds leading zeros
                save_name = str(poke_num).zfill(3) + ' ' + poke_name + file_ext
                if end_url == "Generation_VI_menu_sprites" and not os.path.exists(gen6_menu_sprite_save_path + save_name):
                    print("DNE")
                    #filename, headers = opener.retrieve(get_largest_png(img), gen6_menu_sprite_save_path + save_name)
                if end_url == "Generation_VIII_menu_sprites" and not os.path.exists(gen8_menu_sprite_save_path + save_name):
                    print("DNE")
                    #filename, headers = opener.retrieve(get_largest_png(img), gen8_menu_sprite_save_path + save_name)
            
            try:
                next_page_url = curr_ms_page_soup.find('a', string='next page').get('href')
                next_page = requests.get("https://archives.bulbagarden.net/" + next_page_url)
                next_page_soup = BeautifulSoup(next_page.content, 'html.parser')
                curr_ms_page_soup = next_page_soup
                theres_a_next_page = True
                print("Reading next page of menu sprite archive links...")
            # Unless the end of the next pages is reached
            except:
                theres_a_next_page = False
                print("Reached end of menu sprite archive links.")

# SPREADSHEET DATA
pokemon_info = load_workbook(filename = 'C:\\Users\\ejone\\OneDrive\\Desktop\\Code\\Javascript\\p5\\projects\\Pokeball Pokemon Comparison\\Pokemon Info.xlsx', data_only=True)
pokemon_info_sheet = pokemon_info.worksheets[0]
pokemon_files = load_workbook(filename = 'C:\\Users\\ejone\\OneDrive\\Desktop\\Code\\Javascript\\p5\\projects\\Pokeball Pokemon Comparison\\Pokemon File-check.xlsx', data_only=True)
pokemon_files_sheet = pokemon_files.worksheets[0]

def cell_value(row, col, sheet):
    return (sheet.cell(row, col).value)

def isnt_empty(row, col, sheet):
    return (cell_value(row, col, sheet) != None)

def is_empty(row, col, sheet):
    return (cell_value(row, col, sheet) == None)

# Returns column number from column name
def get_col_number(col_name, sheet):
    for col in range(1, sheet.max_column):
        if (cell_value(1, col, sheet) == col_name):
            return col

# Returns column name from column number
def get_col_name(col_number, sheet):
    return(cell_value(1, col_number, sheet))


def check_if_animated(link):
    # NOTE: Works on animated pngs
    # Converting URL image to PIL Image Object
    # NOTE: May have to use a mask on this requests.get
    img = Image.open(requests.get(link, stream = True).raw)
    # Checking if it is an animated image
    return(img.is_animated)


def get_largest_png(img):
    # If multiple sized images, grab the largest
    try:
        # Sourceset is a string of urls seperated by a comma
            # This breaks that into a list and takes the last (and largest) png url
        srcset = img['srcset'].split(",")
        src = srcset[len(srcset) - 1]
    # If there's only one photo, theres no srcset and that's the largest
    except:
        src = img['src']

    return (src)

def get_img_from_string(img, s, save_path):
    img_text = img.attrs['alt']
    if re.search(s, img_text) != None:
        save_img = get_largest_png(img)
        file_ext = img_text[len(img_text) - 4:]
        print(img_text)
        print(s, " --- ", save_path + file_ext)
        #filename, headers = opener.retrieve(save_img, save_path + file_ext)

# Determines if a pokemon can only be obtained in SM-USUM (so exclude XY-ORAS in filename)
def sm_usum_exclusivity_test(poke_num, tags):
    if poke_num >= 722:
        return True
    if "-Region-Alola" in tags:
        return True
    # For Cap Pikachu
    if "-Form-Cap" in tags:
        return True
    # For Ash Greninja
    if "-Form-Ash" in tags:
        return True
    # For Zygarde
    if "-Form-10%" in tags or "-Form-Complete" in tags:
        return True

def combine_gen_and_game(game, poke_num, tags):
    if game == "Red-Blue" or game == "Red-Green" or game == "Yellow":
        return ("Gen1 " + game)
    if game == "Crystal" or game == "Gold" or game == "Silver":
        return ("Gen2 " + game)
    if game == "Emerald" or game == "FireRed-LeafGreen" or game == "Ruby-Sapphire":
        return ("Gen3 " + game)
    if game == "Diamond-Pearl" or game == "HGSS" or game == "Platinum":
        return ("Gen4 " + game)
    if game == "BW-B2W2":
        return ("Gen5 " + game)
    # Since XY-ORAS and SM-USUM shared sprites
    if game == "XY-ORAS" or game == "SM-USUM":
        is_sm_usum_exclusive = sm_usum_exclusivity_test(poke_num, tags)
        if is_sm_usum_exclusive:
            return ("Gen7 " + game)
        else:
            return ("Gen6-7 XY-ORAS-SM-USUM")
    if game == "LGPE":
        return ("Gen7 " + game)
    if game == "Sword-Shield":
        return ("Gen8 " + game)

# Get back gen from game
    # Handles more sophistocated cases -- Regions, forms, etc
def get_back_gen(game, poke_num, tags):
    if game == "Red-Blue" or game == "Red-Green" or game == "Yellow":
        return ("Gen1")
    if game == "Crystal" or game == "Gold" or game == "Silver":
        return ("Gen2")
    if game == "Emerald" or game == "FireRed-LeafGreen" or game == "Ruby-Sapphire":
        return ("Gen3")
    if game == "Diamond-Pearl" or game == "HGSS" or game == "Platinum":
        return ("Gen4")
    if game == "BW-B2W2":
        return ("Gen5")
    # Since XY-ORAS and SM-USUM shared sprites
    if game == "XY-ORAS" or game == "SM-USUM":
        is_sm_usum_exclusive = sm_usum_exclusivity_test(poke_num, tags)
        if is_sm_usum_exclusive:
            return ("Gen7")
        else:
            return ("Gen6-7")
    if game == "LGPE":
        return ("Gen7")
    if game == "Sword-Shield":
        return ("Gen8")

# Determines what gen to start reading games from
    # Returns gen - 1 for what it's respective index is in the gen array
    # Only goes up to 4 because that's where the potential back discrepancies go up to
def get_back_gen_index_starter(poke_num):
    if poke_num <= 151:
        return (0)
    if poke_num >= 152 and poke_num <= 251:
        return (1)
    if poke_num >= 252 and poke_num <= 386:
        return (2)
    if poke_num >= 387 and poke_num <= 493:
        return (3)

def bulba_game_denoter_conversion(filename):
    if "Red-Blue" in filename:
        return (" 1b")
    if "Red-Green" in filename:
        return (" 1g")
    if "Yellow" in filename:
        return (" 1y")
    if "Crystal" in filename:
        return (" 2c")
    # This has the Gen2 requirement to protect from Golduck and Goldeen entering this conditional by default
    if "Gold" in filename and "Gen2" in filename:
        return (" 2g")
    if "Silver" in filename:
        return (" 2s")
    if "Emerald" in filename:
        return (" 3e")
    if "FireRed-LeafGreen" in filename:
        return (" 3f")
    if "Ruby-Sapphire" in filename:
        return (" 3r")
    if "Diamond-Pearl" in filename:
        return (" 4d")
    if "HGSS" in filename:
        return (" 4h")
    if "Platinum" in filename:
        return (" 4p")
    # 5b2 is converted to 5b if the filename contains 5b2
        # aka Gen5 Black2/White2 (in bulba) is converted to 5b where it will be deemed BW-B2W2 (in my files)
    if "BW-B2W2" in filename:
        return (" 5b")
    # 6o is converted to 6x if the filename contains 6o
        # aka Gen6 ORAS (in bulba) is converted to 6x where it will be deemed XY-ORAS-SM-USUM (in my files)
    if "XY-ORAS" in filename:
        return (" 6x")
    if "SM-USUM" in filename:
        return (" 7s")
    if "LGPE" in filename:
        return (" 7p")
    if "Sword-Shield" in filename:
        return (" 8s")

def check_for_form(plaintext_form, bulba_form_code, curr_bulba_form, computer_filename):
    if plaintext_form in computer_filename:
        return(bulba_form_code)
    else:
        return (curr_bulba_form)

def check_for_type(computer_filename):
    for t in types:
        poke_type = "-Form-" + t
        if poke_type in computer_filename:
            # Handling the ??? type used in gen4
            if t == "Qmark":
                return("-Unknown")
            else:
                return("-" + t)

uppers = list(string.ascii_uppercase)
types = ["Normal", "Fighting", "Flying", "Poison", "Ground", "Rock", "Bug", "Ghost", "Steel", "Fire", "Water", "Grass", "Electric", "Psychic", "Ice", "Dragon", "Dark", "Fairy", "Qmark"]
# Vanilla Cream is default, so no letter denoter
creams = [("Caramel_Swirl", "CaS"), ("Lemon_Cream", "LeC"), ("Matcha_Cream", "MaC"), ("Mint_Cream", "MiC"), ("Rainbow_Swirl", "RaS"), ("Ruby_Cream", "RaC"), ("Ruby_Swirl", "RuS"), ("Salted_Cream", "SaC"), ("Vanilla_Cream", "")]
# Strawberry Sweet is default, so no letter denoter
sweets = [("Berry_Sweet", "B"), ("Clover_Sweet", "C"), ("Flower_Sweet", "F"), ("Love_Sweet", "L"), ("Ribbon_Sweet", "R"), ("Star_Sweet", "S"), ("Strawberry_Sweet", "")]

# Converts forms into bulbapedia notation
def form_translation(pokemon, computer_filename):
    # NOTE: Probably should've done this in like a dict, then just checked keys then tuple arrays... Oops
    bulba_code_form = ""
    # If pokemon has no type or misc forms, return empty string to concatonate onto bulba filename
        # This is running before any web scraping, so I don't need to intentionally slow the script down
    if not pokemon.has_misc_forms and not pokemon.has_type_forms:
        return(bulba_code_form)

    # Pikachu Cosplay & Caps
    if pokemon.name == "Pikachu":
        bulba_code_form = check_for_form("-Form-Cap-Alola", "A", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Cap-Hoenn", "H", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Cap-Kalos", "K", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Cap-Original", "O", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Cap-Sinnoh", "S", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Cap-Unova", "U", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Cap-Partner", "P", bulba_code_form, computer_filename)
        # NOTE: No world cap sprite
        # NOTE: No Sprites for Cosplay on bulbapedia
        # get_img_from_string(img, "^\d\d\dPikachu-Belle.png", drawn_save_path + save_name + "-Cosplay-Belle")
        # get_img_from_string(img, "^\d\d\dPikachu-Libre.png", drawn_save_path + save_name + "-Cosplay-Libre")
        # get_img_from_string(img, "^\d\d\dPikachu-PhD.png", drawn_save_path + save_name + "-Cosplay-PhD")
        # get_img_from_string(img, "^\d\d\dPikachu-Pop Star.png", drawn_save_path + save_name + "-Cosplay-Pop_Star")
        # get_img_from_string(img, "^\d\d\dPikachu-Rock Star.png", drawn_save_path + save_name + "-Cosplay-Rock_Star")

    # Spiky-eared Pichu
    if pokemon.name == "Pichu":
        bulba_code_form = check_for_form("-Form-Spiky_Eared", "N", bulba_code_form, computer_filename)

    # Unown Characters
    # NOTE: Some gens, character is right after pokemon number but on other gens (see gen 4...) there is a hyphen -- UGH
    if pokemon.name == "Unown":
        # Regular letters
        for letter in uppers:
            unown_form = "-Form-" + letter
            if unown_form in computer_filename:
                bulba_code_form = letter
                break
        # Exclamation and question mark
        if "-Form-!" in computer_filename:
            bulba_code_form = "EX"    
        if "-Form-Qmark" in computer_filename:
            bulba_code_form = "QU"


    # Castform Weathers
    if pokemon.name == "Castform":
        bulba_code_form = check_for_form("-Form-Rainy", "R", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Snowy", "H", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Sunny", "S", bulba_code_form, computer_filename)

    # Primal Kyogre & Groudon
    if pokemon.name == "Kyogre" or pokemon.name == "Groudon":
        bulba_code_form = check_for_form("-Form-Primal", "P", bulba_code_form, computer_filename)

    # Deoxys
    if pokemon.name == "Deoxys":
        bulba_code_form = check_for_form("-Form-Attack", "A", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Defense", "D", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Speed", "S", bulba_code_form, computer_filename)

    # Burmy & Wormadam Cloaks
    if pokemon.name == "Burmy" or pokemon.name == "Wormadam":
        # Plant Cloak considered default, so does not have a letter denoter
        bulba_code_form = check_for_form("-Form-Plant_Cloak", "", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Sandy_Cloak", "G", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Trash_Cloak", "S", bulba_code_form, computer_filename)

    # Cherrim
    if pokemon.name == "Cherrim":
        # Overcast form considered default, so does not have a letter denoter
        bulba_code_form = check_for_form("-Form-Overcast", "", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Sunshine", "S", bulba_code_form, computer_filename)

    # Shellos & Gastrodon East/West
    if pokemon.name == "Shellos" or pokemon.name == "Gastrodon":
        # West form considered default, so does not have a letter denoter
        bulba_code_form = check_for_form("-Form-West", "", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-East", "E", bulba_code_form, computer_filename)

    # Rotom Appliances
    if pokemon.name == "Rotom":
        bulba_code_form = check_for_form("-Form-Fan", "F", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Frost", "R", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Heat", "O", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Mow", "L", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Wash", "W", bulba_code_form, computer_filename)

    # Giratina
    if pokemon.name == "Giratina":
        # Altered form considered default, so does not have a letter denoter
        bulba_code_form = check_for_form("-Form-Altered", "", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Origin", "O", bulba_code_form, computer_filename)

    # Shaymin
    if pokemon.name == "Shaymin":
        # Land form considered default, so does not have a letter denoter
        bulba_code_form = check_for_form("-Form-Land", "", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Sky", "S", bulba_code_form, computer_filename)

    # Arceus Types
    if pokemon.name == "Arceus":
        bulba_code_form = check_for_type(computer_filename)

    # Basculin Stripes
    if pokemon.name == "Basculin":
        # Red Striped form considered default, so does not have a letter denoter
        bulba_code_form = check_for_form("-Form-Red_Striped", "", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Blue_Striped", "B", bulba_code_form, computer_filename)

    # Darmanitan Modes
    if pokemon.name == "Darmanitan":
        # Standard form considered default, so does not have a letter denoter
        bulba_code_form = check_for_form("-Form-Standard", "", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Zen", "Z", bulba_code_form, computer_filename)

    # Deerling & Sawsbuck Seasons
    if pokemon.name == "Deerling" or pokemon.name == "Sawsbuck":
        # Spring form considered default, so does not have a letter denoter
        bulba_code_form = check_for_form("-Form-Spring", "", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Autumn", "A", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Summer", "S", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Winter", "W", bulba_code_form, computer_filename)

    # Forces of nature forms
    if pokemon.name == "Tornadus" or pokemon.name == "Thundurus" or pokemon.name == "Landorus":
        # Incarnate form considered default, so does not have a letter denoter
        bulba_code_form = check_for_form("-Form-Incarnate", "", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Therian", "T", bulba_code_form, computer_filename)

    # Kyurem Fusions
    # NOTE: Overdrives were mislabelled as defaults, so I did these by hand
    
    # Keldeo
    if pokemon.name == "Keldeo":
        # Ordinary form considered default, so does not have a letter denoter
        bulba_code_form = check_for_form("-Form-Ordinary", "", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Resolute", "R", bulba_code_form, computer_filename)

    # Meloetta
    if pokemon.name == "Meloetta":
        bulba_code_form = check_for_form("-Form-Aria", "", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Pirouette", "P", bulba_code_form, computer_filename)

    # Genesect
    if pokemon.name == "Genesect":
        if "-Form-Douse_Drive" in computer_filename:
            bulba_code_form = "B"
        if "-Form-Burn_Drive" in computer_filename:
            bulba_code_form = "R"
        if "-Form-Chill_Drive" in computer_filename:
            bulba_code_form = "W"
        if "-Form-Shock_Drive" in computer_filename:
            bulba_code_form = "Y"

    # Ash Greninja
    if pokemon.name == "Greninja":
        bulba_code_form = check_for_form("-Form-Ash", "A", bulba_code_form, computer_filename)

    # Vivillon Patterns
    if pokemon.name == "Vivillon":
        # Meadow form considered default, so does not have a letter denoter
        bulba_code_form = check_for_form("-Form-Meadow", "", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Archipelago", "Arc", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Continental", "Con", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Elegant", "Ele", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Garden", "Gar", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-High_Plains", "Hig", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Icy_Snow", "Icy", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Jungle", "Jun", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Marine", "Mar", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Modern", "Mod", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Monsoon", "Mon", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Ocean", "Oce", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Polar", "Pol", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-River", "Riv", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Sandstorm", "San", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Savanna", "Sav", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Sun", "Sun", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Tundra", "Tun", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Poke_Ball", "Pok", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Fancy", "Fan", bulba_code_form, computer_filename)
    
    # Flabebe, Floette, and Florges colors
    if pokemon.name == "Flabebe" or pokemon.name == "Floette" or pokemon.name == "Florges":
        # Red Flower form considered default, so does not have a letter denoter
        bulba_code_form = check_for_form("-Form-Red_Flower", "", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Blue_Flower", "B", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Orange_Flower", "O", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-White_Flower", "W", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Yellow_Flower", "Y", bulba_code_form, computer_filename)

    # Furfrou Trims
    if pokemon.name == "Furfrou":
        bulba_code_form = check_for_form("-Form-Dandy_Trim", "Da", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Debutante_Trim", "De", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Diamond_Trim", "Di", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Heart_Trim", "He", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Kabuki_Trim", "Ka", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-La_Reine_Trim", "La", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Matron_Trim", "Ma", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Pharaoh_Trim", "Ph", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Star_Trim", "St", bulba_code_form, computer_filename)

    # Aegislash
    if pokemon.name == "Aegislash":
        # Shield form considered default, so does not have a letter denoter
        bulba_code_form = check_for_form("-Form-Shield", "", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Blade", "B", bulba_code_form, computer_filename)

    # Pumpkaboo and Gourgeist Sizes
    if pokemon.name == "Pumpkaboo" or pokemon.name == "Gourgeist":
        # Average Size form considered default, so does not have a letter denoter
        bulba_code_form = check_for_form("-Form-1_Average_Size", "", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-0_Small_Size", "Sm", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-2_Large_Size", "La", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-3_Super_Size", "Su", bulba_code_form, computer_filename)

    # Xerneas
    if pokemon.name == "Xerneas":
        # Active form considered default, so does not have a letter denoter
        bulba_code_form = check_for_form("-Form-Active", "", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Neutral", "N", bulba_code_form, computer_filename)

    # Zygarde
    if pokemon.name == "Zygarde":
        # 50% form considered default, so does not have a letter denoter
        bulba_code_form = check_for_form("-Form-50%", "", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Complete", "C", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-10%", "T", bulba_code_form, computer_filename)

    # Hoopa
    if pokemon.name == "Hoopa":
        # Confined form considered default, so does not have a letter denoter
        bulba_code_form = check_for_form("-Form-Confined", "", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Unbound", "U", bulba_code_form, computer_filename)

    # Oricorio
    if pokemon.name == "Oricorio":
        # Confined form considered default, so does not have a letter denoter
        bulba_code_form = check_for_form("-Form-Baile", "", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Pa'u", "Pa", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Pom_Pom", "Po", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Sensu", "Se", bulba_code_form, computer_filename)

    # Lycanroc
    if pokemon.name == "Lycanroc":
        # Midday form considered default, so does not have a letter denoter
        bulba_code_form = check_for_form("-Form-Midday", "", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Dusk", "D", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Midnight", "Mn", bulba_code_form, computer_filename)

    # Wishiwashi
    if pokemon.name == "Wishiwashi":
        # Solo form considered default, so does not have a letter denoter
        bulba_code_form = check_for_form("-Form-Solo", "", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-School", "Sc", bulba_code_form, computer_filename)

    # Silvally Types
    if pokemon.name == "Silvally":
        bulba_code_form = check_for_type(computer_filename)

    # Minior
    # NOTE: Bulba has shiny core form denoted as red, so this will probably have to manually be changed
    if pokemon.name == "Minior":
        # Meteor form considered default, so does not have a letter denoter
        bulba_code_form = check_for_form("-Form-Meteor", "", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Blue_Core", "B", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Green_Core", "G", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Indigo_Core", "I", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Orange_Core", "O", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Red_Core", "R", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Violet_Core", "V", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Yellow_Core", "Y", bulba_code_form, computer_filename)

    # Mimikyu
    if pokemon.name == "Mimikyu":
        # Disguised form considered default, so does not have a letter denoter
        bulba_code_form = check_for_form("-Form-Disguised", "", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Busted", "B", bulba_code_form, computer_filename)

    # Solgaleo
    if pokemon.name == "Solgaleo":
        bulba_code_form = check_for_form("-Form-Full_Moon", "F", bulba_code_form, computer_filename)

    # Lunala
    if pokemon.name == "Lunala":
        bulba_code_form = check_for_form("-Form-Radiant_Sun", "R", bulba_code_form, computer_filename)

    # Necrozma
    if pokemon.name == "Necrozma":
        bulba_code_form = check_for_form("-Form-Dawn_Wings", "DW", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Dusk_Mane", "DM", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Ultra", "U", bulba_code_form, computer_filename)

    # Magearna
    if pokemon.name == "Magearna":
        bulba_code_form = check_for_form("-Form-Original_Color", "O", bulba_code_form, computer_filename)
    
    # Marshadow
    # NOTE: Bulba does not have Zenith form
    #if pokemon.name == "Marshadow":
    #    bulba_code_form = check_for_form("-Form-Zenith", "Z", bulba_code_form, computer_filename)
    
    # Cramorant
    if pokemon.name == "Cramorant":
        bulba_code_form = check_for_form("-Form-Gorging", "Go", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Gulping", "Gu", bulba_code_form, computer_filename)
    
    # Toxtricity
    if pokemon.name == "Toxtricity":
        # Amped form considered default, so does not have a letter denoter
        bulba_code_form = check_for_form("-Form-Amped", "", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Low_Key", "L", bulba_code_form, computer_filename)

    # Alcremie Creams & Sweets
    if pokemon.name == "Alcremie":
        # Gigantamax version doesn't show cream or sweets, so there's no point to go through them
        if "-Form-Gigantamax" in computer_filename:
            return("")
            
        for sweet in sweets:
            if sweet[0] in computer_filename:
                # Shiny Alcremie only shows sweet, not cream color
                if "-Shiny" in computer_filename:
                    bulba_code_form = sweet[1]
                    break
                else:
                    # If regular color, find cream
                    for cream in creams:
                        if cream[0] in computer_filename:
                            bulba_code_form = cream[1] + sweet[1]
                            break

    # Eiscue
    if pokemon.name == "Eiscue":
        # Ice Face form considered default, so does not have a letter denoter
        bulba_code_form = check_for_form("-Form-Ice_Face", "", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Noice_Face", "N", bulba_code_form, computer_filename)
    
    # Morpeko
    if pokemon.name == "Morpeko":
        # Full Belly form considered default, so does not have a letter denoter
        bulba_code_form = check_for_form("-Form-Full_Belly", "", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Hangry", "H", bulba_code_form, computer_filename)

    # Zacian
    if pokemon.name == "Zacian":
        # Hero form considered default, so does not have a letter denoter
        bulba_code_form = check_for_form("-Form-Hero_of_Many_Battles", "", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Crowned_Sword", "C", bulba_code_form, computer_filename)

    # Zamazenta
    if pokemon.name == "Zamazenta":
        # Hero form considered default, so does not have a letter denoter
        bulba_code_form = check_for_form("-Form-Hero_of_Many_Battles", "", bulba_code_form, computer_filename)
        bulba_code_form = check_for_form("-Form-Crowned_Shield", "C", bulba_code_form, computer_filename)

    # Eternatus Eternamax
    if pokemon.name == "Eternatus":
        bulba_code_form = check_for_form("-Form-Eternamax", "E", bulba_code_form, computer_filename)

    # Urshifu
    # NOTE: NO game sprites for Urshifu??!

    # Zarude
    # NOTE: NO game sprites for Zarude either??!

    # Calyrex Ridings
    # NOTE: And no game sprites for Calyrex...

    return(bulba_code_form)

 no_bulba_forms = []
# Pikachu World Cap
no_bulba_forms.append("-Form-Cap-World")
# Cosplay Pikachu
no_bulba_forms.append("-Form-Cosplay")
# Overdrive Reshiram, Zekrom, and Kyurem
no_bulba_forms.append("Overdrive")
# Marshadow Zenith
no_bulba_forms.append("-Form-Zenith")
# Urshifu Forms
no_bulba_forms.extend(["-Form-Rapid_Strike", "-Form-Single_Strike"])
# Dada Zarude
no_bulba_forms.append("-Form-Dada")
# Calyrex Riders
no_bulba_forms.extend(["-Form-Shadow_Rider", "-Form-Ice_Rider"])
def bulba_doesnt_have_this_form(filename):
    for form in no_bulba_forms:
        if form in filename:
            return True

    return False
    

exception_strings = []
# Mega
exception_strings.append("M")
# Gigantamax
exception_strings.append("Gi")
# Regions
exception_strings.extend(["A", "G"])
# Pikachu caps
# NOTE: To be honest this shouldn't even be needed since the first check is for a -f in the filename, which the cap variants don't have
    # But what the hell, err on the side of caution
exception_strings.extend(["025A", "025H", "025K", "025O", "025S", "025U", "025P"])

# Checks if there's a string in the bulba filename (usually from a form)
    # That excepts the need of a male denoter (m) 
        # This is then passed to where this denoter is added, recognizing if it should be or not
def has_male_denoter_exception(bulba_filename):
    for ex_str in exception_strings:
        if ex_str in bulba_filename:
            return True
    return False

# Converts my filename structure to bulbapedias
def determine_bulba_name(computer_filename, pokemon):
    # All files start with Spr
    bulba_name = "Spr"
    # Back denotions first
    is_gen1_back = False
    if "-Back" in computer_filename:
        bulba_name += " b"
        if " Gen1" in computer_filename:
            is_gen1_back = True
            bulba_name += " g1"
    # Then Game denoters
    if not is_gen1_back:
        bulba_name += bulba_game_denoter_conversion(computer_filename)
    # Then pokedex number
    bulba_name += " " + computer_filename[:3]
    
    # Then Mega
        # Not gender specific, so can go before gender check
    if "-Mega" in computer_filename:
        bulba_name += "M"

    # Then Region
        # Not gender specific, so can go before gender check
    if "-Region-Alola" in computer_filename:
        bulba_name += "A"
    if "-Region-Galar" in computer_filename:
        bulba_name += "G"

    # Then Forms
        # MUST COME AFTER REGION
            # See Darmanitan (555)
    bulba_name += form_translation(pokemon, computer_filename)

    # Then Gigantamax
        # MUST COME AFTER FORMS (see Urshifu 892)
            # But not gender specific, so can go before gender check
    if "-Gigantamax" in computer_filename:
        bulba_name += "Gi"

    # Then gender
    if "-f" in computer_filename:
        bulba_name += " f"
    else:
        # Bulbapedia puts m denoters into filenames for male version
            # So if the female denoter is missing in my filename, but the species has a gender difference
                # Check for this and exceptions and add the male denoter if needed
        has_m_exception = has_male_denoter_exception(bulba_name)
        if pokemon.has_f_var and not has_m_exception:
            bulba_name += " m"

    # Then shiny
    if "Shiny" in computer_filename:
        bulba_name += " s"

    return (bulba_name)

    
# Pokemon object
class Pokemon:
    def __init__(self, name, number, gen, has_f_var, has_mega, has_giganta, reg_forms, has_type_forms, has_misc_forms, is_in_gen8):
        self.name = name
        self.number = number
        self.gen = gen
        self.has_f_var = has_f_var
        self.has_mega = has_mega
        self.has_giganta = has_giganta
        self.reg_forms = reg_forms
        self.has_type_forms = has_type_forms
        self.has_misc_forms = has_misc_forms
        self.is_in_gen8 = is_in_gen8
        self.missing_imgs = []
        self.missing_gen1_thru_gen4_back_imgs = []

# Gets column numbers from spreadsheet
name_col = get_col_number("Name", pokemon_info_sheet)
num_col = get_col_number("#", pokemon_info_sheet)
gen_col = get_col_number("Gen", pokemon_info_sheet)
f_col = get_col_number("Female Variation", pokemon_info_sheet)
mega_col = get_col_number("Mega", pokemon_info_sheet)
giganta_col = get_col_number("Gigantamax", pokemon_info_sheet)
reg_forms_col = get_col_number("Regional Forms", pokemon_info_sheet)
type_forms_col = get_col_number("Type Forms", pokemon_info_sheet)
misc_forms_col = get_col_number("Misc Forms", pokemon_info_sheet)
gen8_col = get_col_number("Available in Gen 8", pokemon_info_sheet)

# Adds pokemon info from spreadsheet to object array
print("Getting pokemon info from spreadsheet...")
pokedex = []
for i in range(2, 900):
    name = cell_value(i, name_col, pokemon_info_sheet)
    num = cell_value(i, num_col, pokemon_info_sheet)
    gen = int(cell_value(i, gen_col, pokemon_info_sheet))
    has_f_var = isnt_empty(i, f_col, pokemon_info_sheet)
    has_mega = isnt_empty(i, mega_col, pokemon_info_sheet)
    has_giganta = isnt_empty(i, giganta_col, pokemon_info_sheet)
    reg_forms = cell_value(i, reg_forms_col, pokemon_info_sheet)
    has_type_forms = isnt_empty(i, type_forms_col, pokemon_info_sheet)
    has_misc_forms = isnt_empty(i, misc_forms_col, pokemon_info_sheet)
    is_in_gen8 = isnt_empty(i, gen8_col, pokemon_info_sheet)

    pokedex.append(Pokemon(name, num, gen, has_f_var, has_mega, has_giganta, reg_forms, has_type_forms, has_misc_forms, is_in_gen8))

# Getting missing images
print("Getting missing images from spreadsheet...")
missing_imgs = {}
poke_num_col = get_col_number("#", pokemon_files_sheet)
poke_name_col = get_col_number("Name", pokemon_files_sheet)
tags_col = get_col_number("Tags", pokemon_files_sheet)
filename_col = get_col_number("Filename", pokemon_files_sheet)
# These are for back generation differences to be able to loop through and concatonate game name to filename
gen1_games = ["Yellow", "Red-Green", "Red-Blue"]
gen2_games = ["Silver", "Gold", "Crystal"]
gen3_games = ["Ruby-Sapphire", "FireRed-LeafGreen", "Emerald"]
gen4_games = ["Platinum", "HGSS", "Diamond-Pearl"]
gen_1_thru_4_games = [gen1_games, gen2_games, gen3_games, gen4_games]

for row in range(2, pokemon_files_sheet.max_row):
    poke_num = int(cell_value(row, poke_num_col, pokemon_files_sheet))
    poke_name = cell_value(row, poke_name_col, pokemon_files_sheet)
    poke_obj = -1
    for pokemon in pokedex:
        if pokemon.name == poke_name:
            poke_obj = pokemon
    tags = cell_value(row, tags_col, pokemon_files_sheet)
    if tags == None:
        tags = ""
    filename = cell_value(row, filename_col, pokemon_files_sheet)

    if bulba_doesnt_have_this_form(filename):
        continue

    # This is to track generations and skip if it's a back sprite below gen 5
        # Those sprites are being pulled seperately in the row only loop above
            # To be sifted through to see if they're different by game for the given pokemon
    is_below_gen5 = False
    # Only doing filename_col up because those are where the actual checks need to be made (missing for certain games)
        # And +1 at the end to be inclusive
    for col in range(filename_col + 1, pokemon_files_sheet.max_column + 1):
        # Triggers at Platinum because exxcel file is reverse chronological, so Plat is first gen 4 game hit
            # Every loop iteration after is_below_gen5 will be true
        if cell_value(1, col, pokemon_files_sheet) == "Platinum":
            is_below_gen5 = True
        
        # If pokemon image is unavailable, continue
        if cell_value(row, col, pokemon_files_sheet) == "u":
            continue

        col_name = get_col_name(col, pokemon_files_sheet)

        # If it's a back image from a pokemon between gen1 and gen4
            # Put all game images into special missing array
                # This is so another script can go through these images and determine if there were differences in the sprites between games
                    # If there were, each file will be named differently
                    # Otherwise, they will all be lumped into a single gen# back img
        is_back_below_gen5 = is_below_gen5 and "-Back" in filename
        if is_empty(row, col, pokemon_files_sheet) or is_back_below_gen5:
            gen_insert_index = filename.find(poke_name) + len(poke_name)
            gen_and_game = combine_gen_and_game(col_name, poke_num, tags)
            
            # Back sprites have to be seperated due to:
                # Below gen 5 all back sprites are being downloaded to a seperate folder to be sifted through
            if "-Back" in tags:
                # Adding space because initially there was no space in spreadsheet
                    # This was for sorting purposes because back sprites start with hyphen immediately after gen
                        # But front sprites follow with a space so they come first in file order
                # This variable is necessary because bulba uses games to denote which back sprites are from where
                    # In my filenaming convention (gen4 and below excluded due to actual sprite differences between games)
                        # For backs I JUST use gen
                            # But to scrape the image from bulba, I still need the game denoter, which this gets me
                back_filename_w_game_and_gen_for_bulba = filename[:gen_insert_index] + " " + gen_and_game + filename[gen_insert_index:]
                bulba_name = determine_bulba_name(back_filename_w_game_and_gen_for_bulba, poke_obj)
                actual_filename =""
                gen = ""
                game = ""
                if "Gen6" in gen_and_game:
                    # This is because all Gen6 are shared with gen7
                        # So denoter is Gen6-7
                    gen = gen_and_game[:6]
                    # Extra character to skip over space after gen
                    game = gen_and_game[7:]
                else:
                    gen = gen_and_game[:4]
                    # Extra character to skip over space after gen
                    game = gen_and_game[5:]

                # filename[:3] gets pokemon number
                actual_filename = filename[:3] + " " + poke_name + " " + gen + filename[gen_insert_index:]
                if is_below_gen5:
                    actual_filename += "-" + game
                    poke_obj.missing_gen1_thru_gen4_back_imgs.append((bulba_name, actual_filename))
                else:
                    poke_obj.missing_imgs.append((bulba_name, actual_filename))
                #print(actual_filename, "     changed to     ", bulba_name)
            else:
                # Going +1 after the insert index because there's a space for non-back sprites
                # This is to simulate in the spreadsheet the space between generation and games in the filenames
                    # This determines sorting order, so is fairly important
                        # Since the spreadsheet doesn't acknowledge games or gens in the "Filename" column (because each game gets it's own column)
                            # The space must be included to sort the spreadsheet to match filenames
                                # And thus, must be accounted for here
                gen_insert_index += 1
                filename_w_gen = filename[:gen_insert_index] + gen_and_game + filename[gen_insert_index:]
                bulba_name = determine_bulba_name(filename_w_gen, poke_obj)
                poke_obj.missing_imgs.append((bulba_name, filename_w_gen))
                
                #print(filename_w_gen, "     changed to     ", bulba_name)
    # print(poke_obj.name)
    # for img in poke_obj.missing_imgs:
    #     print(img)
    # print("\n\n")

# Origin page (list of pokes by national pokedex)
starter_url = "https://archives.bulbagarden.net"
pokemon_starter_page = requests.get("https://archives.bulbagarden.net/wiki/Category:Pok%C3%A9mon_artwork")
pokemon_starter_page_soup = BeautifulSoup(pokemon_starter_page.content, 'html.parser')
save_path_starter = "C:\\Users\\ejone\\OneDrive\\Desktop\\Code\\Javascript\\p5\\projects\\Pokeball Pokemon Comparison\\Images\\Pokemon"
drawn_save_path = save_path_starter + "\\Drawn\\"
game_save_path = save_path_starter + "\\Game Sprites\\"
gen1_thru_4_backs_save_path = game_save_path + "\\missing_back_imgs_to_be_filtered\\"
gen6_menu_sprite_save_path = save_path_starter + "\\Menu Sprites\\Gen6\\"
gen8_menu_sprite_save_path = save_path_starter + "\\Menu Sprites\\Gen8\\"

get_menu_sprites()

pokemon_img_urls = []
curr_page_soup = pokemon_starter_page_soup
print("Starting reading of pokemon game sprite archive links...")

page_index = 0
# Loops through pages of archives of pokemon images
while True:
    # Stopping after a certain page for testing
    # if page_index == 2:
    #     break

    # Grabbing each individual pokemons archived image url
    for list_div in curr_page_soup.find_all('div', {'class': 'mw-category-group'}):
        for poke in list_div.find_all('li'):
            # Skipping specific artwork I don't want
            if poke.a.get('href') == "/wiki/Category:Ken_Sugimori_Pok%C3%A9mon_artwork" or poke.a.get('href') == "/wiki/Category:Official_Pok%C3%A9mon_artwork":
                continue
            pokemon_img_urls.append(poke.a.get('href'))

    # TODO: Remove this before running
    # Only gets first page of pokemon archive links
    break

    # Moving on to the next page
    try:
        next_page_url = curr_page_soup.find('a', string='next page').get('href')
        next_page = requests.get(starter_url + next_page_url)
        next_page_soup = BeautifulSoup(next_page.content, 'html.parser')
        curr_page_soup = next_page_soup
        page_index += 1
        print("Reading next page of pokemon archive links...")
    # Unless the end of the next pages is reached
    except:
        print("Reached end of pokemon archive links.")
        break

# If the file contains lateral generation changes (black to black2, xy to oras) change the filename to accomodate my script
def potentially_adapt_game_in_filename(filename):
    if " 5b2 " in filename:
        return (filename.replace(" 5b2 ", " 5b "))
    if " 6o " in filename:
        return (filename.replace(" 6o ", " 6x "))

imgs_downloaded = 0
imgs_still_missing = []
print("Processing game sprite images...")
for i in range(len(pokemon_img_urls)):
    # Getting relevant pokemon data
    pokemon = pokedex[i]
    # Converting to dicts so I can search if an image name is in the dict using keyword in
    missing_imgs = dict(pokemon.missing_imgs)
    missing_gen1_thru_gen4_back_imgs = dict(pokemon.missing_gen1_thru_gen4_back_imgs)
    print(pokemon.name, " has ", len(missing_imgs) + len(missing_gen1_thru_gen4_back_imgs), " missing images...")
    # Fixing a bulba error here the adds a hyphen between unown number and form in gen4
    if pokemon.name == "Unown":
        for old_k, v in missing_imgs.items():
            if "Gen4" in v:
                new_k = old_k.replace(" 201", " 201-")
                missing_imgs[new_k] = missing_imgs.pop(old_k)
        for old_k, v in missing_gen1_thru_gen4_back_imgs.items():
            if "Gen4" in v:
                new_k = old_k.replace(" 201", " 201-")
                missing_imgs[new_k] = missing_gen1_thru_gen4_back_imgs.pop(old_k)

    # For only doing certain pokemon
    # if pokemon.name != "Alcremie":
    #     continue
    # print("Got here")

    # Getting pokemon archived image page information
    curr_page = requests.get(starter_url + pokemon_img_urls[i])
    curr_page_soup = BeautifulSoup(curr_page.content, 'html.parser')

    theres_a_next_page = True
    theres_more_imgs = True
    while (theres_a_next_page and theres_more_imgs):
        pokemon_imgs = curr_page_soup.find_all('img')
        # Downloading certain images
        for img in pokemon_imgs:
            img_text = img.attrs['alt']
            # Only run drawn images if it starts with 3 digits and then pokemon name
            if re.search("^\d\d\d[a-zA-Z]", img_text) != None:
                get_drawn_images(pokemon, img)
            # Skipping image if it's not a sprite image
            if not img_text.startswith("Spr"):
                continue
            # Goes to next pokemon if there's no more images to gather
            if len(missing_imgs) == 0 and len(missing_gen1_thru_gen4_back_imgs) == 0:
                # Necessary to break out of while loop
                theres_more_imgs = False
                break
            
            img_text_wo_file_ext = img_text[:len(img_text)-4]
            file_ext = img_text[len(img_text) - 4:]
            save_name = missing_imgs[img_text_wo_file_ext] + file_ext
            # Change 5b to 5b2 and 6x to 6o if needed to grab those sprites
            img_text_wo_file_ext = potentially_adapt_game_in_filename(img_text_wo_file_ext)
            

            # All missing images
            if img_text_wo_file_ext in missing_imgs:
                if "-Animated" in save_name:
                    if check_if_animated(get_largest_png(img)):
                        # filename, headers = opener.retrieve(get_largest_png(img), game_save_path + "animated_pngs_for_gifs\\pngs\\" + save_name)
                        # Removing missing image from list
                        dummy = missing_imgs.pop(img_text, None)
                        imgs_downloaded += 1
                        continue
                    else:
                        print(save_name, " was not animated... Skipped")
                        continue
                else:
                    # Making sure its NOT animated
                    if not check_if_animated(get_largest_png(img)):
                        # filename, headers = opener.retrieve(get_largest_png(img), game_save_path + "initial_downloads_for_border_removal\\" + save_name)
                        # Removing missing image from list
                        dummy = missing_imgs.pop(img_text, None)
                        imgs_downloaded += 1
                        continue
                    else:
                        # If it is animated, still download it with an obvious denoter to convert it to a static
                        # filename, headers = opener.retrieve(get_largest_png(img), game_save_path + "initial_downloads_for_border_removal\\" + "TO_BE_CONVERTED_TO_STILL_" + save_name)
                        # Removing missing image from list
                        dummy = missing_imgs.pop(img_text, None)
                        imgs_downloaded += 1
                        continue
                
            # ONLY Gen1 thru Gen4 back sprites to test for difference
            if img_text_wo_file_ext in missing_gen1_thru_gen4_back_imgs:
                if "-Animated" in save_name:
                    if check_if_animated(get_largest_png(img)):
                        # filename, headers = opener.retrieve(get_largest_png(img), game_save_path + "back_imgs_to_be_filtered\\animated\\" + save_name)
                        # Removing missing image from list
                        dummy = missing_gen1_thru_gen4_back_imgs.pop(img_text, None)
                        imgs_downloaded += 1
                        continue
                    else:
                        print(save_name, " was not animated... Skipped")
                        continue
                else:
                    # Making sure its NOT animated
                    if not check_if_animated(get_largest_png(img)):
                        # filename, headers = opener.retrieve(get_largest_png(img), game_save_path + "back_imgs_to_be_filtered\\static\\" + save_name)
                        # Removing missing image from list
                        dummy = missing_gen1_thru_gen4_back_imgs.pop(img_text, None)
                        imgs_downloaded += 1
                        continue
                    else:
                        # If it is animated, still download it with an obvious denoter to convert it to a still
                        # filename, headers = opener.retrieve(get_largest_png(img), game_save_path + "back_imgs_to_be_filtered\\static\\" + "TO_BE_CONVERTED_TO_STILL_" + save_name)
                        # Removing missing image from list
                        dummy = missing_gen1_thru_gen4_back_imgs.pop(img_text, None)
                        imgs_downloaded += 1
                        continue


            # TODO: READ ALL TODOS BEFORE YOU RUN THIS -- SOME OF THEM ARE VERY IMPORTANT!!!
            # TODO: Before running, uncomment all filename, headers = opener.retrieve(get_largest_png(img), gen8_menu_sprite_save_path + save_name)

        # Moving on to the next page
        try:
            next_page_url = curr_page_soup.find('a', string='next page').get('href')
            next_page = requests.get(starter_url + next_page_url)
            next_page_soup = BeautifulSoup(next_page.content, 'html.parser')
            curr_page_soup = next_page_soup
            theres_a_next_page = True
            print("Reading next page of ", pokemon.name, " archive links...")
        # Unless the end of the next pages is reached
        except:
            theres_a_next_page = False
            print("Reached end of ", pokemon.name, " archive links.")
            break

    # To see how many images I am still missing
    for k,v in missing_imgs.items():
        imgs_still_missing.append(v)
    for k,v in missing_gen1_thru_gen4_back_imgs.items():
        imgs_still_missing.append(v)

for i in imgs_still_missing:
    print(i)
print (imgs_downloaded,  " images downloaded")
print (len(imgs_still_missing),  " images still missing (see above)")
    